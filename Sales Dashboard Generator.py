import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.drawing.image import Image
import os
from datetime import datetime

class SalesDashboardGenerator:
    def __init__(self, data_path):
        self.data = pd.read_csv(data_path)
        self.report_date = datetime.now().strftime("%Y-%m-%d")
        self.output_dir = "reports"
        os.makedirs(self.output_dir, exist_ok=True)

    def clean_data(self):
        """Clean and preprocess sales data"""
        # Convert date column to datetime
        self.data['Date'] = pd.to_datetime(self.data['Date'])
        
        # Calculate additional metrics
        self.data['Total_Sales'] = self.data['Quantity'] * self.data['Unit_Price']
        self.data['Month'] = self.data['Date'].dt.month_name()
        
        # Handle missing values
        self.data.fillna({'Region': 'Unknown'}, inplace=True)
        
        return self.data

    def generate_visualizations(self):
        """Create matplotlib visualizations"""
        plt.style.use('ggplot')
        
        # Monthly Sales Trend
        monthly_sales = self.data.groupby('Month')['Total_Sales'].sum()
        plt.figure(figsize=(10, 5))
        monthly_sales.plot(kind='line', marker='o', color='royalblue')
        plt.title('Monthly Sales Trend', fontweight='bold')
        plt.ylabel('Total Sales ($)')
        plt.savefig(f"{self.output_dir}/monthly_trend.png")
        plt.close()
        
        # Top Products
        top_products = self.data.groupby('Product')['Total_Sales'].sum().nlargest(5)
        plt.figure(figsize=(8, 8))
        top_products.plot(kind='pie', autopct='%1.1f%%')
        plt.title('Top Selling Products', fontweight='bold')
        plt.ylabel('')
        plt.savefig(f"{self.output_dir}/top_products.png")
        plt.close()
        
        # Regional Performance
        region_sales = self.data.groupby('Region')['Total_Sales'].sum()
        plt.figure(figsize=(10, 5))
        region_sales.plot(kind='bar', color='forestgreen')
        plt.title('Sales by Region', fontweight='bold')
        plt.ylabel('Total Sales ($)')
        plt.xticks(rotation=45)
        plt.savefig(f"{self.output_dir}/regional_sales.png")
        plt.close()

    def generate_excel_report(self):
        """Generate formatted Excel report with visuals"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Sales Summary"
        
        # Add header
        ws['A1'] = f"Sales Dashboard Report - {self.report_date}"
        ws['A1'].font = Font(bold=True, size=14)
        
        # Add visualizations
        img_paths = [
            f"{self.output_dir}/monthly_trend.png",
            f"{self.output_dir}/top_products.png", 
            f"{self.output_dir}/regional_sales.png"
        ]
        
        for i, img_path in enumerate(img_paths, start=3):
            img = Image(img_path)
            img.width, img.height = 500, 300
            ws.add_image(img, f"A{i}")
        
        # Add data summary
        summary_row = len(img_paths) + 5
        ws[f"A{summary_row}"] = "Key Metrics"
        ws[f"A{summary_row}"].font = Font(bold=True)
        
        metrics = [
            ("Total Sales", self.data['Total_Sales'].sum()),
            ("Average Order Value", self.data['Total_Sales'].mean()),
            ("Top Product", self.data.groupby('Product')['Total_Sales'].sum().idxmax()),
            ("Best Region", self.data.groupby('Region')['Total_Sales'].sum().idxmax())
        ]
        
        for i, (label, value) in enumerate(metrics, start=summary_row + 1):
            ws[f"A{i}"] = label
            ws[f"B{i}"] = value if isinstance(value, str) else f"${value:,.2f}"
        
        # Save Excel file
        report_path = f"{self.output_dir}/Sales_Dashboard_{self.report_date}.xlsx"
        wb.save(report_path)
        return report_path

    def generate_pdf_report(self):
        """Generate PDF report (placeholder - would use libraries like FPDF)"""
        # Implementation would go here
        pass

    def run(self):
        """Execute full pipeline"""
        self.clean_data()
        self.generate_visualizations()
        report_path = self.generate_excel_report()
        print(f"Report generated: {report_path}")

# Sample Usage
if __name__ == "__main__":
    # Example data structure (would normally come from CSV/database)
    sample_data = """Date,Product,Region,Quantity,Unit_Price
2023-01-01,Widget A,North,10,19.99
2023-01-05,Widget B,South,5,29.99
2023-01-10,Widget A,East,8,19.99
2023-02-15,Widget C,West,12,9.99
2023-02-20,Widget B,North,6,29.99
2023-03-05,Widget A,South,15,19.99"""
    
    # Save sample data to CSV
    with open("sample_sales.csv", "w") as f:
        f.write(sample_data)
    
    # Generate report
    dashboard = SalesDashboardGenerator("sample_sales.csv")
    dashboard.run()