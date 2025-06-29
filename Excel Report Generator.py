import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import random

def generate_sample_data():
    products = ['Laptop', 'Phone', 'Tablet', 'Monitor', 'Keyboard']
    regions = ['North', 'South', 'East', 'West']
    
    data = {
        'Date': [datetime(2023, random.randint(1, 12), random.randint(1, 28)).date() 
                 for _ in range(100)],
        'Product': [random.choice(products) for _ in range(100)],
        'Region': [random.choice(regions) for _ in range(100)],
        'Units_Sold': [random.randint(1, 50) for _ in range(100)],
        'Unit_Price': [round(random.uniform(100, 2000), 2) for _ in range(100)]
    }
    
    df = pd.DataFrame(data)
    df['Total_Sales'] = df['Units_Sold'] * df['Unit_Price']
    return df

def create_excel_report(df, output_file):

    
    # Create summary statistics
    summary = df.groupby('Product').agg({
        'Units_Sold': 'sum',
        'Total_Sales': ['sum', 'mean']
    }).reset_index()
    summary.columns = ['Product', 'Total Units Sold', 'Total Revenue', 'Average Sale']
    
    # Create a pivot table
    pivot = pd.pivot_table(df, 
                         values='Total_Sales',
                         index='Product',
                         columns='Region',
                         aggfunc='sum',
                         fill_value=0)
    
    # Create Excel workbook
    wb = Workbook()
    
    # Remove default sheet and create our own
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
   # Create sheets
    ws_data = wb.create_sheet('Raw Data')
    ws_summary = wb.create_sheet('Summary')
    ws_pivot = wb.create_sheet('Pivot Table')
    
    # Write data to sheets
    for r in dataframe_to_rows(df, index=False, header=True):
        ws_data.append(r)
    
    for r in dataframe_to_rows(summary, index=False, header=True):
        ws_summary.append(r)
    
    for r in dataframe_to_rows(pivot, index=True, header=True):
        ws_pivot.append(r)
    
    # Apply formatting
    format_worksheet(ws_data, 'Sales Raw Data')
    format_worksheet(ws_summary, 'Sales Summary')
    format_worksheet(ws_pivot, 'Sales by Region')
    
    # Save the workbook
    wb.save(output_file)
    print(f"Report generated successfully: {output_file}")

def format_worksheet(ws, title):
    
    # Set column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width
    
    # Header formatting
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='thin'), 
                        bottom=Side(style='thin'))
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    
    # Add title
    ws.insert_rows(1)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(ws[2]))
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = Font(size=14, bold=True)
    title_cell.alignment = Alignment(horizontal='center')
    
    # Format numbers
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                if cell.value > 1000:  # Assume currency
                    cell.number_format = '$#,##0.00'
                else:
                    cell.number_format = '0'
                cell.alignment = Alignment(horizontal='right')

def main():
    # Generate sample data
    sales_data = generate_sample_data()
    
    # Create report
    output_filename = f"Sales_Report_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    create_excel_report(sales_data, output_filename)

if __name__ == "__main__":
    main()
